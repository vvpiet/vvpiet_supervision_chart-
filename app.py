import streamlit as st
from scheduler import generate_exam_dates, generate_schedule, build_supervisor_table
from pdf_utils import generate_duty_pdf, combine_pdfs_bytes, generate_combined_duty_pdf
import inspect
import importlib


def _call_pdf_compat(func, supervisor_name, schedule_df, staff_df, start_date, end_date, exam_type, college_logo_bytes=None, uni_logo_bytes=None, sign_bytes=None):
    """Call a PDF function (generate_duty_pdf or generate_combined_duty_pdf) with only the optional kwargs it supports.
    This avoids TypeError when older deployed versions of pdf_utils have fewer parameters.
    """
    sig = inspect.signature(func)
    supported = {}
    if 'college_logo_bytes' in sig.parameters:
        supported['college_logo_bytes'] = college_logo_bytes
    if 'uni_logo_bytes' in sig.parameters:
        supported['uni_logo_bytes'] = uni_logo_bytes
    if 'sign_bytes' in sig.parameters:
        supported['sign_bytes'] = sign_bytes
    return func(supervisor_name, schedule_df, staff_df, start_date, end_date, exam_type, **supported)


def _call_memo_compat(func, supervisor_name, absences, staff_df, college_logo_bytes=None, uni_logo_bytes=None, sign_bytes=None):
    """Call a memo PDF function with only the optional kwargs it supports (backwards compatible)."""
    sig = inspect.signature(func)
    supported = {}
    if 'college_logo_bytes' in sig.parameters:
        supported['college_logo_bytes'] = college_logo_bytes
    if 'uni_logo_bytes' in sig.parameters:
        supported['uni_logo_bytes'] = uni_logo_bytes
    if 'sign_bytes' in sig.parameters:
        supported['sign_bytes'] = sign_bytes
    return func(supervisor_name, absences, staff_df, **supported)
from email_utils import send_email_with_attachment
import pandas as pd
import io
import datetime
import os
import subprocess
import sys
import json
import ast


def save_attendance_state(att_map):
    try:
        with open("attendance_state.json", "w", encoding="utf-8") as f:
            json.dump(att_map, f, ensure_ascii=False)
    except Exception:
        st.warning("Unable to persist attendance state to disk; attendance may be lost on refresh.")


def load_attendance_state():
    if os.path.exists("attendance_state.json"):
        try:
            with open("attendance_state.json", "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_schedule_state(df, meta: dict):
    try:
        # Convert dates to ISO strings so JSON is safe
        temp = df.copy()
        if 'date' in temp.columns:
            temp['date'] = temp['date'].apply(lambda d: d.isoformat() if not pd.isnull(d) else None)
        records = temp.to_dict(orient='records')
        payload = {'meta': meta, 'records': records}
        with open('schedule_state.json', 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False)
        # Also save a pickle for more robust reloading (preserves lists, dtypes)
        try:
            df.to_pickle('schedule_state.pkl')
            with open('schedule_meta.json', 'w', encoding='utf-8') as mf:
                json.dump(meta, mf, ensure_ascii=False)
        except Exception:
            # non-fatal, JSON remains
            pass
        # Also write a CSV backup for easy manual restore/download
        try:
            df.to_csv('schedule_state.csv', index=False)
        except Exception:
            pass
    except Exception:
        st.warning('Unable to persist schedule to disk; schedule may be lost on refresh.')


def load_schedule_state():
    # Prefer pickle (more robust), fall back to JSON
    if os.path.exists('schedule_state.pkl'):
        try:
            df = pd.read_pickle('schedule_state.pkl')
            meta = {}
            try:
                with open('schedule_meta.json', 'r', encoding='utf-8') as mf:
                    meta = json.load(mf)
            except Exception:
                meta = {}
            return df, meta
        except Exception:
            # Try JSON fallback
            pass

    if os.path.exists('schedule_state.json'):
        try:
            with open('schedule_state.json', 'r', encoding='utf-8') as f:
                payload = json.load(f)
            records = payload.get('records', [])
            if not records:
                return None, None
            df = pd.DataFrame(records)
            if 'date' in df.columns:
                df['date'] = pd.to_datetime(df['date']).dt.date
            meta = payload.get('meta', {})
            return df, meta
        except Exception:
            return None, None
    return None, None


def ensure_schedule_schema(df):
    """Ensure schedule DataFrame has required columns: 'date', 'session', 'assigned'.
    Normalize column names (case-insensitive), convert dates to date objects and parse assigned lists when stored as strings.
    Returns normalized df or None if it cannot be normalized.
    """
    if df is None or df.empty:
        return None

    df = df.copy()

    # Normalize column names (case-insensitive mapping)
    cols = {c.lower(): c for c in df.columns}
    # date
    if 'date' not in cols:
        # try common variants
        for candidate in ['Date', 'DATE']:
            if candidate in df.columns:
                df = df.rename(columns={candidate: 'date'})
                cols['date'] = 'date'
                break
        # try any column whose lower() == 'date'
        if 'date' not in cols:
            for c in df.columns:
                if c.strip().lower() == 'date':
                    df = df.rename(columns={c: 'date'})
                    cols['date'] = 'date'
                    break

    if 'date' in cols:
        try:
            df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.date
        except Exception:
            pass
    else:
        return None

    # session
    if 'session' not in cols:
        for c in df.columns:
            if c.strip().lower() == 'session':
                df = df.rename(columns={c: 'session'})
                cols['session'] = 'session'
                break
    if 'session' not in cols:
        # If session column missing, try to infer based on 'time' or create default
        if 'time' in (c.lower() for c in df.columns):
            for c in df.columns:
                if c.lower() == 'time':
                    df = df.rename(columns={c: 'session'})
                    cols['session'] = 'session'
                    break
        else:
            # Can't reliably infer session; bail
            return None

    # assigned
    if 'assigned' not in cols:
        for c in df.columns:
            if c.strip().lower() == 'assigned':
                df = df.rename(columns={c: 'assigned'})
                cols['assigned'] = 'assigned'
                break
    if 'assigned' not in cols:
        # maybe assignments were stored in a column like 'Assigned' or 'assigned_to'
        for c in df.columns:
            if 'assign' in c.strip().lower():
                df = df.rename(columns={c: 'assigned'})
                cols['assigned'] = 'assigned'
                break

    if 'assigned' in cols:
        # Ensure assigned are lists; parse string reprs
        def _parse_assigned(v):
            if v is None:
                return []
            if isinstance(v, list):
                return v
            if isinstance(v, (float, int)) and pd.isna(v):
                return []
            if isinstance(v, str):
                s = v.strip()
                # try literal_eval for list-like strings
                try:
                    val = ast.literal_eval(s)
                    if isinstance(val, list):
                        return [str(x) for x in val]
                except Exception:
                    # fallback: comma separated
                    if ',' in s:
                        return [x.strip() for x in s.split(',') if x.strip()]
                    if s == '':
                        return []
                    return [s]
            return [v]

        df['assigned'] = df['assigned'].apply(_parse_assigned)
    else:
        return None

    return df


def _map_common_schedule_columns(df):
    """Try to map common column name variants to the expected schema keys."""
    if df is None or df.empty:
        return df
    rename = {}
    for c in df.columns:
        lc = c.strip().lower()
        if 'date' in lc or 'exam date' in lc or 'date of' in lc:
            rename[c] = 'date'
        elif lc in ('session', 'time', 'shift') or 'session' in lc or 'time' in lc or 'shift' in lc:
            rename[c] = 'session'
        elif 'assign' in lc or 'invigil' in lc or 'supervisor' in lc or 'name' in lc:
            # careful: avoid renaming 'Name of Faculty' to 'assigned' incorrectly when schedule horizontal format is used; only map if column seems to contain multiple names
            rename[c] = 'assigned'
    if rename:
        try:
            return df.rename(columns=rename)
        except Exception:
            return df
    return df


def _resolve_duplicate_columns(df):
    """Resolve duplicate column names by merging or selecting sensible values.
    - 'assigned': combine multiple columns into a single list of names
    - 'date': pick first non-null and coerce to date
    - 'session': pick first non-null
    For any other duplicated column, pick the first non-null value.
    """
    import collections
    if df is None or df.empty:
        return df

    # First make duplicate column names unique by appending a suffix so we can address them individually
    cols = list(df.columns)
    seen = {}
    unique_cols = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            unique_cols.append(f"{c}__dup{seen[c]}")
        else:
            seen[c] = 0
            unique_cols.append(c)
    tmp = df.copy()
    tmp.columns = unique_cols

    # Group by base lowercased names (strip duplication suffix)
    groups = collections.defaultdict(list)
    for orig in tmp.columns:
        base = orig.split('__dup')[0].strip().lower()
        groups[base].append(orig)

    out = tmp.copy()
    for low, originals in groups.items():
        if len(originals) <= 1:
            continue
        # Merge depending on key
        if low == 'assigned':
            def merge_assigned(row):
                vals = []
                for c in originals:
                    v = row.get(c)
                    if pd.isna(v) or v is None:
                        continue
                    # reuse ensure_schedule_schema parsing logic (simple):
                    if isinstance(v, list):
                        vals.extend([str(x) for x in v])
                    else:
                        s = str(v).strip()
                        if s.startswith('[') and s.endswith(']'):
                            try:
                                arr = ast.literal_eval(s)
                                if isinstance(arr, list):
                                    vals.extend([str(x) for x in arr])
                                    continue
                            except Exception:
                                pass
                        if ',' in s:
                            vals.extend([x.strip() for x in s.split(',') if x.strip()])
                        elif s:
                            vals.append(s)
                # unique preserve order
                seen = set()
                out_list = []
                for x in vals:
                    if x not in seen:
                        seen.add(x)
                        out_list.append(x)
                return out_list

            out['assigned'] = out.apply(merge_assigned, axis=1)
        elif low == 'date':
            # pick first non-null and coerce to date
            def pick_date(row):
                for c in originals:
                    v = row.get(c)
                    if pd.isna(v) or v is None or str(v).strip() == '':
                        continue
                    try:
                        return pd.to_datetime(v).date()
                    except Exception:
                        continue
                return None

            out['date'] = out.apply(pick_date, axis=1)
        elif low == 'session':
            def pick_session(row):
                for c in originals:
                    v = row.get(c)
                    if pd.isna(v) or v is None or str(v).strip() == '':
                        continue
                    return str(v).strip()
                return None

            out['session'] = out.apply(pick_session, axis=1)
        else:
            # generic: pick first non-null
            def pick_first(row):
                for c in originals:
                    v = row.get(c)
                    if pd.isna(v) or v is None or str(v).strip() == '':
                        continue
                    return v
                return None

            out[low] = out.apply(pick_first, axis=1)

        # drop other original columns except the standardized one we created
        for c in originals:
            if c != low and c in out.columns:
                try:
                    out.drop(columns=[c], inplace=True)
                except Exception:
                    pass

    return out

st.set_page_config(page_title="Exam Supervision Allotment", layout="wide")

st.title("Supervision Allotment and Duty Orders")

st.sidebar.header("Staff / Upload")
uploaded = st.sidebar.file_uploader("Upload staff CSV (must contain Name and Mail Id)", type=["csv"] )

# Persist uploaded staff CSV so it survives refreshes on the server
PERSISTED_STAFF = os.path.join(os.getcwd(), "staff_uploaded.csv")
if uploaded is not None:
    try:
        # read bytes and persist to disk
        b = uploaded.getvalue()
    except Exception:
        try:
            uploaded.seek(0)
            b = uploaded.read()
        except Exception:
            b = None

    if b:
        try:
            with open(PERSISTED_STAFF, "wb") as f:
                f.write(b)
        except Exception:
            st.warning("Unable to persist uploaded staff CSV to disk; data may be lost on refresh.")

    try:
        staff_df = pd.read_csv(io.BytesIO(b), header=0)
    except Exception:
        try:
            staff_df = pd.read_csv(uploaded, header=0)
        except Exception:
            staff_df = pd.DataFrame(columns=["Sr. No.", "Name of Supervisor", "Mail Id"])
else:
    # If there is a previously uploaded staff file on disk, prefer it
    if os.path.exists(PERSISTED_STAFF):
        try:
            staff_df = pd.read_csv(PERSISTED_STAFF, header=0)
        except Exception:
            staff_df = pd.DataFrame(columns=["Sr. No.", "Name of Supervisor", "Mail Id"])
    else:
        default_path = os.path.join(os.getcwd(), "Staff List Uniform list (1).csv")
        try:
            staff_df = pd.read_csv(default_path, header=0)
        except Exception:
            staff_df = pd.DataFrame(columns=["Sr. No.", "Name of Supervisor", "Mail Id"])

st.sidebar.write(f"Loaded {len(staff_df)} supervisors")
st.sidebar.info("Uploaded staff CSV is persisted to the app storage as 'staff_uploaded.csv' and attendance is auto-saved to 'attendance_state.json' so refresh won't lose data.")
if st.sidebar.button("Clear persisted schedule"):
    try:
        if os.path.exists('schedule_state.json'):
            os.remove('schedule_state.json')
        if 'schedule_df' in st.session_state:
            del st.session_state['schedule_df']
        if 'schedule_meta' in st.session_state:
            del st.session_state['schedule_meta']
        st.sidebar.success('Persisted schedule cleared. You may need to re-generate the schedule.')
    except Exception:
        st.sidebar.error('Unable to remove persisted schedule file.')

# Restore schedule from a local CSV backup if needed
with st.sidebar.expander("Restore schedule from file"):
    restore_file = st.file_uploader("Upload a schedule CSV to restore", type=["csv"], key="restore_schedule")
    if restore_file is not None:
        try:
            df_restore = pd.read_csv(restore_file, header=0)
            df_restore = _map_common_schedule_columns(df_restore)
            df_restore = _resolve_duplicate_columns(df_restore)
            # try to parse date column
            if 'date' in df_restore.columns:
                try:
                    df_restore['date'] = pd.to_datetime(df_restore['date']).dt.date
                except Exception:
                    pass
            norm = ensure_schedule_schema(df_restore)
            if norm is None:
                st.sidebar.error('Uploaded schedule appears malformed: required columns (date, session, assigned) not found or could not be parsed. Please fix the CSV and try again.')
            else:
                st.session_state['schedule_df'] = norm
                save_schedule_state(norm, {'restored_from_upload': True, 'restored_at': datetime.datetime.now().isoformat()})
                st.sidebar.success('Schedule restored into session and persisted.')
        except Exception as e:
            st.sidebar.error(f'Unable to restore schedule from uploaded file: {e}')

st.header("Exam Configuration")
col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Start of Examinations", datetime.date.today())
    end_date = st.date_input("End of Examinations", datetime.date.today() + datetime.timedelta(days=7))
    exam_type = st.selectbox("Exam Type", ["Regular", "Supplementary"])
with col2:
    exclude_weekends = st.checkbox("Skip Sundays only", value=True, help="When checked, exam dates will skip Sundays only (Saturday will be included).")
    holiday_text = st.text_area("Holidays (comma separated YYYY-MM-DD)", help="Enter dates separated by commas")
    try:
        holidays = [datetime.datetime.strptime(d.strip(), "%Y-%m-%d").date() for d in holiday_text.split(",") if d.strip()]
    except Exception:
        holidays = []
    st.markdown("---")
    st.subheader("SMTP Configuration (for sending emails)")
    smtp_server = st.text_input("SMTP server", value=st.session_state.get("smtp_server", "smtp.gmail.com"))
    smtp_port = st.text_input("SMTP port", value=st.session_state.get("smtp_port", "587"))
    smtp_user = st.text_input("SMTP user (From email)", value=st.session_state.get("smtp_user", ""))
    smtp_password = st.text_input("SMTP password (app password recommended)", type="password", value=st.session_state.get("smtp_password", ""))
    if st.button("Save SMTP settings"):
        st.session_state["smtp_server"] = smtp_server
        st.session_state["smtp_port"] = smtp_port
        st.session_state["smtp_user"] = smtp_user
        st.session_state["smtp_password"] = smtp_password
        st.success("SMTP settings saved in session")
    if st.button("Test SMTP connection"):
        import smtplib
        try:
            with smtplib.SMTP(smtp_server, int(smtp_port)) as smtp:
                smtp.starttls()
                if smtp_user and smtp_password:
                    smtp.login(smtp_user, smtp_password)
            st.success("SMTP connection successful")
        except Exception as e:
            st.error(f"SMTP test failed: {e}")

st.markdown("---")
st.subheader("Week-wise Day-wise Block Configuration")
st.info("Configure the number of blocks for each day in each week of the exam period.")

# Calculate exam dates and organize by weeks
try:
    exam_dates = generate_exam_dates(start_date, end_date, exclude_weekends, holidays)
    
    if len(exam_dates) == 0:
        st.warning("No exam dates found. Please check your date range and holiday settings.")
        day_blocks = {}
    else:
        # Organize exam dates by weeks
        # A week starts on Monday (weekday 0)
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        
        # Group dates by ISO calendar week
        weeks_dict = {}  # Format: {(year, week_number): [list of dates in that week]}
        for d in exam_dates:
            iso_year, iso_week, iso_weekday = d.isocalendar()
            week_key = (iso_year, iso_week)
            if week_key not in weeks_dict:
                weeks_dict[week_key] = []
            weeks_dict[week_key].append(d)
        
        # Sort weeks chronologically
        sorted_weeks = sorted(weeks_dict.items())
        
        # Display summary
        st.markdown("### Exam Period Summary")
        st.markdown(f"**Total Exam Days:** {len(exam_dates)} | **Total Weeks:** {len(sorted_weeks)} | **Date Range:** {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}")
        if exclude_weekends:
            st.caption("ℹ️ Sundays are excluded from the exam period")
        if holidays:
            st.caption(f"ℹ️ {len(holidays)} holiday(s) excluded from exam period")
        
        st.markdown("---")
        
        # Create week-wise configuration
        day_blocks = {}
        
        for week_idx, (week_key, week_dates) in enumerate(sorted_weeks, 1):
            iso_year, iso_week = week_key
            week_start = min(week_dates)
            week_end = max(week_dates)
            
            st.markdown(f"### Week {week_idx} ({week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')})")
            
            # Count which days are in this week
            week_day_count = {day: 0 for day in day_names}
            for d in week_dates:
                day_name = day_names[d.weekday()]
                week_day_count[day_name] += 1
            
            # Show metrics for days present in this week
            summary_cols = st.columns(min(6, sum(1 for c in week_day_count.values() if c > 0)))
            col_idx = 0
            for day in day_names[:6]:  # Only Monday-Saturday
                if week_day_count[day] > 0:
                    with summary_cols[col_idx]:
                        st.metric(label=day, value=f"{week_day_count[day]} day(s)")
                    col_idx += 1
            
            # Input fields for this week's days
            st.markdown(f"**Set blocks for Week {week_idx}:**")
            week_cols = st.columns(2)
            col_idx = 0
            
            for day in day_names[:6]:  # Only Monday-Saturday
                if week_day_count[day] > 0:
                    with week_cols[col_idx % 2]:
                        st.markdown(f"**{day}**")
                        morning_val = st.number_input(
                            f"Morning blocks", 
                            min_value=1, 
                            max_value=30, 
                            value=2,
                            key=f"week_{week_idx}_{day}_morning"
                        )
                        evening_val = st.number_input(
                            f"Evening blocks", 
                            min_value=1, 
                            max_value=30, 
                            value=2,
                            key=f"week_{week_idx}_{day}_evening"
                        )
                        # Store with week prefix so different weeks can have different configs
                        day_blocks[f"week_{week_idx}_{day}"] = {
                            "morning": morning_val,
                            "evening": evening_val,
                            "week_number": week_idx
                        }
                    col_idx += 1
            
            st.markdown("---")

except Exception as e:
    st.error(f"Unable to calculate exam dates: {e}")
    exam_dates = []
    day_blocks = {}

st.caption("📝 The supervision chart will be generated based on these week-wise day-wise configurations for your exam period.")

# Special blocks: per-date overrides (only for Supplementary exams)
special_blocks = {}
special_date_session_blocks = {}  # Format: {date: {"morning": blocks, "evening": blocks}}

if exam_type == "Supplementary":
    st.markdown("---")
    st.info("Optionally define specific dates that override the day-wise configuration.")
    st.markdown("**Option 1: Simple (same blocks for both sessions on a date)**")
    st.caption("Format: YYYY-MM-DD:blocks (one per line) - applies same block count to both Morning and Evening")
    special_input = st.text_area("Special dates with blocks (format YYYY-MM-DD:blocks)", height=100, key="special_blocks_simple")
    for line in special_input.splitlines():
        if ":" in line:
            d, b = line.split(":")
            try:
                sd = datetime.datetime.strptime(d.strip(), "%Y-%m-%d").date()
                special_blocks[sd] = int(b.strip())
            except Exception:
                pass
    
    st.markdown("**Option 2: Detailed (different blocks for Morning and Evening)**")
    st.caption("Format: YYYY-MM-DD:Morning:blocks,Evening:blocks (one per line) - Example: 2026-01-22:Morning:3,Evening:2")
    detailed_input = st.text_area("Special dates with session-wise blocks", height=120, key="special_blocks_detailed")
    for line in detailed_input.splitlines():
        if ":" in line and "Morning" in line and "Evening" in line:
            parts = line.split(":")
            try:
                sd = datetime.datetime.strptime(parts[0].strip(), "%Y-%m-%d").date()
                # Parse Morning:blocks and Evening:blocks
                session_data = {}
                for i in range(1, len(parts), 2):
                    session_name = parts[i].strip().lower()
                    if i + 1 < len(parts):
                        block_count = int(parts[i+1].strip().rstrip(','))
                        session_data[session_name] = block_count
                if session_data:
                    special_date_session_blocks[sd] = session_data
            except Exception:
                pass

st.subheader("College/University Logos (Optional)")
col_logo = st.file_uploader("College logo (left)", type=["png","jpg","jpeg"], key="college_logo")
uni_logo = st.file_uploader("University logo (right)", type=["png","jpg","jpeg"], key="uni_logo")
college_logo_bytes = col_logo.read() if col_logo else None
uni_logo_bytes = uni_logo.read() if uni_logo else None

# Try to load a previously generated schedule so attendance is available without regenerating
if 'schedule_df' not in st.session_state:
    loaded_df, loaded_meta = load_schedule_state()
    if loaded_df is not None:
        # Try auto-mapping common column variants before normalization and resolve duplicates
        mapped = _map_common_schedule_columns(loaded_df)
        resolved = _resolve_duplicate_columns(mapped)

        # Determine generated timestamp for staleness check
        ts = None
        try:
            if loaded_meta and 'generated_at' in loaded_meta:
                ts = loaded_meta.get('generated_at')
        except Exception:
            ts = None
        if not ts:
            try:
                ts = datetime.datetime.fromtimestamp(os.path.getmtime('schedule_state.pkl' if os.path.exists('schedule_state.pkl') else 'schedule_state.json')).isoformat()
            except Exception:
                ts = None

        # Stale handling: don't auto-load schedules older than TTL_DAYS unless user confirms
        TTL_DAYS = 7
        stale = False
        if ts:
            try:
                ts_dt = datetime.datetime.fromisoformat(ts)
                age_days = (datetime.datetime.now() - ts_dt).days
                if age_days > TTL_DAYS:
                    stale = True
            except Exception:
                stale = False

        if stale:
            # Store candidate in session and prompt user via sidebar
            st.session_state['schedule_persisted_candidate_df'] = resolved
            st.session_state['schedule_persisted_candidate_meta'] = loaded_meta or {}
            st.session_state['schedule_loaded_from'] = 'disk_stale'
            st.session_state['schedule_loaded_timestamp'] = ts
            st.sidebar.warning(f"A persisted schedule dated {ts[:19]} is older than {TTL_DAYS} days. You can Load it or Clear persisted schedule.")
            if st.sidebar.button('Load persisted schedule (force)', key='load_persisted_force'):
                norm = ensure_schedule_schema(resolved)
                if norm is not None:
                    st.session_state['schedule_df'] = norm
                    st.session_state['schedule_meta'] = loaded_meta or {}
                    st.session_state['schedule_loaded_from'] = 'disk'
                    st.session_state['schedule_loaded_timestamp'] = ts
                    st.sidebar.success('Persisted schedule loaded into session.')
                else:
                    st.sidebar.error('Persisted schedule looks malformed and could not be normalized.')
        else:
            norm = ensure_schedule_schema(resolved)
            if norm is not None:
                st.session_state['schedule_df'] = norm
                st.session_state['schedule_meta'] = loaded_meta or {}
                # mark source and timestamp for UI badge
                st.session_state['schedule_loaded_from'] = 'disk'
                st.session_state['schedule_loaded_timestamp'] = ts
                st.info('Loaded previously generated schedule from disk; Attendance Marking is populated.')
            else:
                st.warning("Found a persisted schedule but it appears malformed (missing required columns). Please restore from a CSV or regenerate the schedule.")
        # make persisted schedule downloadable from sidebar for backup
        try:
            if os.path.exists('schedule_state.csv'):
                with open('schedule_state.csv', 'rb') as f:
                    csv_bytes = f.read()
                st.sidebar.download_button('Download persisted schedule (CSV)', data=csv_bytes, file_name='schedule_state.csv', mime='text/csv')
        except Exception:
            pass

if st.button("Generate Schedule"):
    exam_dates = generate_exam_dates(start_date, end_date, exclude_weekends, holidays)
    
    # Merge special_blocks (simple format) and special_date_session_blocks (detailed format)
    merged_special_blocks = special_blocks.copy()
    
    # Generate schedule based on day-wise block configuration
    # day_blocks: per-day per-session blocks (Monday morning: 2, Monday evening: 2, etc.)
    # special_blocks: per-date same-count overrides (optional, for Supplementary only)
    # date_session_blocks: per-date per-session overrides (optional, for Supplementary only)
    schedule_df = generate_schedule(
        exam_dates, 
        default_blocks=2,  # Fallback if no other config applies
        special_blocks=merged_special_blocks,    # Simple date overrides (both sessions)
        staff_df=staff_df,
        session_blocks=None,  # Not used in day-wise only mode
        day_blocks=day_blocks if day_blocks else None,  # Day-of-week per-session blocks
        date_session_blocks=special_date_session_blocks  # Per-date per-session overrides
    )
    
    st.session_state["schedule_df"] = schedule_df
    # Persist schedule and basic metadata to disk so it can be loaded on refresh
    meta = {
        'start_date': start_date.isoformat() if isinstance(start_date, (datetime.date, datetime.datetime)) else str(start_date),
        'end_date': end_date.isoformat() if isinstance(end_date, (datetime.date, datetime.datetime)) else str(end_date),
        'exam_type': exam_type,
        'day_blocks': day_blocks if day_blocks else {},
        'special_blocks': {d.isoformat(): b for d, b in special_blocks.items()} if special_blocks else {},
        'special_date_session_blocks': {d.isoformat(): session_config for d, session_config in special_date_session_blocks.items()} if special_date_session_blocks else {}
    }
    save_schedule_state(schedule_df, meta)
    # Mark session as freshly generated (so badge shows)
    st.session_state['schedule_loaded_from'] = 'generated'
    st.session_state['schedule_loaded_timestamp'] = datetime.datetime.now().isoformat()
    meta['generated_at'] = st.session_state['schedule_loaded_timestamp']
    # update saved meta with timestamp
    try:
        with open('schedule_meta.json', 'w', encoding='utf-8') as mf:
            json.dump(meta, mf, ensure_ascii=False)
    except Exception:
        pass
    st.success("Schedule generated with day and session-wise block configuration and cached in session.")

if "schedule_df" in st.session_state:
    # Validate and normalize schedule before preview to avoid duplicate-column or missing-column errors
    st.subheader("Schedule Preview")
    preview_df = _resolve_duplicate_columns(st.session_state.get("schedule_df"))
    preview_df = ensure_schedule_schema(preview_df)
    if preview_df is None:
        st.error("Current schedule is malformed. Please restore a valid schedule (use the Restore tool) or regenerate the schedule.")
    else:
        st.dataframe(preview_df)
    # Offer Excel download in required horizontal format
    def schedule_to_excel_bytes(schedule_df):
        # Defensive: normalize/validate incoming DataFrame schema
        safe = ensure_schedule_schema(schedule_df)
        if safe is None:
            # Caller will display an error; return None to indicate failure
            return None
        schedule_df = safe
        # Build workbook with merged headers using openpyxl for precise formatting
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Header rows
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # A1:A2 Sr. No.
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # B1:B2 Name
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)  # C1:D1 Date
        ws.cell(row=1, column=1, value="Sr. No.")
        ws.cell(row=1, column=2, value="Name of Faculty")
        ws.cell(row=1, column=3, value="Date")
        ws.cell(row=2, column=3, value="Morning (10.00 a.m. to 01.00 p.m.)")
        ws.cell(row=2, column=4, value="Evening (02.00 p.m. to 05.00 p.m.)")

        # Style headers
        for r in [1, 2]:
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

        # Fill data grouped by date (date row as merged label, then supervisors)
        row_idx = 3
        sr = 1
        for d in sorted(schedule_df["date"].unique()):
            # write a date separator row merged across columns A:D
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            ws.cell(row=row_idx, column=1, value=d.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1

            morning = schedule_df[(schedule_df["date"] == d) & (schedule_df["session"] == "Morning")]
            evening = schedule_df[(schedule_df["date"] == d) & (schedule_df["session"] == "Evening")]
            morning_assigned = morning.iloc[0]["assigned"] if not morning.empty else []
            evening_assigned = evening.iloc[0]["assigned"] if not evening.empty else []
            supervisors = sorted(set(morning_assigned + evening_assigned))
            for name in supervisors:
                m_tick = "✓" if name in morning_assigned else ""
                e_tick = "✓" if name in evening_assigned else ""
                ws.cell(row=row_idx, column=1, value=sr)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=m_tick)
                ws.cell(row=row_idx, column=4, value=e_tick)
                row_idx += 1
                sr += 1

        # Auto-width columns
        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 25 if col == "B" else 15

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.read()

    excel_bytes = schedule_to_excel_bytes(st.session_state["schedule_df"])
    if excel_bytes is not None:
        filename = f"Schedule_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
        st.download_button("Download Schedule (Excel)", data=excel_bytes, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_schedule_excel")
    # Also provide a horizontal format: rows are supervisors, columns are date-session pairs
    def schedule_to_excel_horizontal(schedule_df):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule_Horizontal"

        # Build list of unique dates
        dates = sorted(schedule_df["date"].unique())
        # Header row: Sr. No., Name, then for each date two columns (Morning, Evening)
        headers = ["Sr. No.", "Name of Faculty"]
        for d in dates:
            headers.append(d.strftime("%Y-%m-%d") + "\nMorning")
            headers.append(d.strftime("%Y-%m-%d") + "\nEvening")

        # Write header
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.font = Font(bold=True)

        # Build list of supervisors
        names = sorted({n for lst in schedule_df["assigned"].tolist() for n in lst})
        row_idx = 2
        sr = 1
        for name in names:
            ws.cell(row=row_idx, column=1, value=sr)
            ws.cell(row=row_idx, column=2, value=name)
            col = 3
            for d in dates:
                morning = schedule_df[(schedule_df["date"] == d) & (schedule_df["session"] == "Morning")]
                evening = schedule_df[(schedule_df["date"] == d) & (schedule_df["session"] == "Evening")]
                morning_assigned = morning.iloc[0]["assigned"] if not morning.empty else []
                evening_assigned = evening.iloc[0]["assigned"] if not evening.empty else []
                ws.cell(row=row_idx, column=col, value="✓" if name in morning_assigned else "")
                ws.cell(row=row_idx, column=col+1, value="✓" if name in evening_assigned else "")
                col += 2
            row_idx += 1
            sr += 1

        # Adjust column widths
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.read()

    # Defensive: ensure schedule has expected schema before generating Excel
    safe_df = ensure_schedule_schema(st.session_state.get("schedule_df"))
    if safe_df is None:
        st.error("Schedule is malformed or missing required columns (e.g., 'date'). Please regenerate or restore a valid schedule.")
    else:
        horiz_bytes = schedule_to_excel_horizontal(safe_df)
        if horiz_bytes is not None:
            filename2 = f"Schedule_Horizontal_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
            st.download_button("Download Schedule (Horizontal Excel)", data=horiz_bytes, file_name=filename2, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_schedule_horizontal")
else:
    st.info("Generate a schedule to preview assignments.")

st.header("Duty Order Download & Email")
# Supervisor selection with select-all/clear
names = staff_df.iloc[:, 1].fillna("Unnamed").tolist()
if "selected_supervisors" not in st.session_state:
    st.session_state["selected_supervisors"] = []
col_a, col_b = st.columns([3, 1])
with col_a:
    sel = st.multiselect("Select supervisor(s) to generate duty allotment", options=names, key="selected_supervisors")
    # Signature for PDFs: check default file sign.jpg, allow upload to override
    default_sign_path = os.path.join(os.getcwd(), "sign.jpg")
    default_sign_bytes = None
    if os.path.exists(default_sign_path):
        try:
            with open(default_sign_path, "rb") as f:
                default_sign_bytes = f.read()
        except Exception:
            default_sign_bytes = None
    sign_file = st.file_uploader("Signature (sign.jpg) for PDFs (optional)", type=["jpg", "jpeg", "png"], key="sign_pdf")
    sign_bytes = sign_file.read() if sign_file else default_sign_bytes
with col_b:
    def _select_all():
        st.session_state["selected_supervisors"] = names

    def _clear_selection():
        st.session_state["selected_supervisors"] = []

    st.button("Select all", on_click=_select_all)
    st.button("Clear", on_click=_clear_selection)


if st.button("Generate & Download PDF for selected"):
    if "schedule_df" not in st.session_state:
        st.error("Please generate schedule first.")
    else:
        schedule_df = st.session_state["schedule_df"]
        pdfs = []
        if not sel:
            st.error("No supervisors selected.")
        else:
            # PDF merger availability check
            def pdf_merger_available():
                try:
                    import pypdf  # noqa: F401
                    return True
                except Exception:
                    try:
                        import PyPDF2  # noqa: F401
                        return True
                    except Exception:
                        return False

            if not pdf_merger_available():
                st.warning("PDF merging libraries not installed. You can install 'pypdf' or 'PyPDF2' now (recommended: pypdf).")
                if st.button("Install pypdf"):
                    # Try multiple python executables to accommodate systems where 'pip' is not on PATH
                    tries = [sys.executable, "python", "py"]
                    success = False
                    install_output = []
                    for exe in tries:
                        try:
                            st.info(f"Trying install with: {exe} -m pip install pypdf")
                            res = subprocess.run([exe, "-m", "pip", "install", "pypdf"], capture_output=True, text=True)
                            install_output.append((exe, res.returncode, res.stdout + '\n' + res.stderr))
                            if res.returncode == 0:
                                st.success(f"pypdf installed successfully with `{exe}`. Please re-run the combined PDF operation.")
                                success = True
                                break
                        except Exception as e:
                            install_output.append((exe, -1, str(e)))
                    if not success:
                        # Show summarized install failure information and recommend manual install
                        st.error("Automatic installation failed. See details and try installing manually with 'python -m pip install pypdf'.")
                        for exe, rc, out in install_output:
                            st.write(f"Attempt with {exe} returned code {rc}. Output:\n{out}")

            for idx, name in enumerate(sel):
                # Use compatibility wrapper to avoid errors if deployed pdf_utils has fewer optional args
                pdf_bytes = _call_pdf_compat(generate_duty_pdf, name, schedule_df, staff_df, start_date, end_date, exam_type, college_logo_bytes, uni_logo_bytes, sign_bytes)
                # Validate PDF has pages before appending
                valid = True
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(io.BytesIO(pdf_bytes))
                    if len(reader.pages) == 0:
                        valid = False
                except Exception:
                    # If pypdf not available, assume valid if bytes non-empty
                    valid = bool(pdf_bytes)

                if not valid:
                    st.warning(f"Generated PDF for {name} appears empty; skipping in combined output.")
                    # Still offer individual download so user can inspect
                    st.download_button(f"Download duty order for {name} (may be empty)", data=pdf_bytes, file_name=f"Duty_{name}.pdf", mime="application/pdf", key=f"download_empty_{idx}_{name}")
                    continue

                pdfs.append(pdf_bytes)
                # Offer individual download
                st.download_button(f"Download duty order for {name}", data=pdf_bytes, file_name=f"Duty_{name}.pdf", mime="application/pdf", key=f"download_{idx}_{name}")
            # If more than one selected then offer combined single PDF
            if len(pdfs) > 1:
                # Prefer direct combined PDF generator (avoids external mergers)
                try:
                    combined = _call_pdf_compat(generate_combined_duty_pdf, sel, schedule_df, staff_df, start_date, end_date, exam_type, college_logo_bytes, uni_logo_bytes, sign_bytes)
                except Exception as gen_e:
                    st.warning(f"Direct combined generator failed ({gen_e}), attempting to merge individual PDFs...")
                    try:
                        combined = combine_pdfs_bytes(pdfs)
                    except Exception as e:
                        st.error(f"Failed to combine PDFs: {e}")
                        combined = None

                if combined:
                    # Count pages if possible
                    page_count = None
                    try:
                        from pypdf import PdfReader
                        reader = PdfReader(io.BytesIO(combined))
                        page_count = len(reader.pages)
                    except Exception:
                        page_count = None
                    st.download_button("Download combined PDF for selected", data=combined, file_name="Combined_Duty_Allotments.pdf", mime="application/pdf", key="download_combined_pdf")
                    if page_count is not None:
                        st.info(f"Combined PDF contains {page_count} pages (one or more pages per faculty as required).")
                        if page_count < len(pdfs):
                            st.warning(f"Combined PDF page count ({page_count}) is less than the number of included PDFs ({len(pdfs)}). Please inspect individual PDFs.")


st.markdown("---")

st.header("Email duty orders")
st.info("Emails are sent using SMTP settings in Streamlit secrets (smtp.server, smtp.port, smtp.user, smtp.password) or environment variables. For Gmail use an app password.")
if st.button("Send emails to selected"):
    if "schedule_df" not in st.session_state:
        st.error("Please generate schedule first.")
    else:
        # Ensure SMTP is configured (session or secrets or env)
        smtp_configured = False
        if st.session_state.get("smtp_server") and st.session_state.get("smtp_user") and st.session_state.get("smtp_password"):
            smtp_configured = True
        else:
            try:
                _ = st.secrets["smtp"]
                smtp_configured = True
            except Exception:
                smtp_configured = False

        if not smtp_configured:
            st.error("SMTP not configured. Set credentials in SMTP Configuration or Streamlit secrets or environment variables before sending emails.")
        else:
            schedule_df = st.session_state["schedule_df"]
            for name in sel:
                matching = staff_df[staff_df.iloc[:, 1].str.strip() == name]
                if matching.empty or matching.shape[0] == 0:
                    st.warning(f"No email for {name}")
                    continue

                # Robustly find an email-like value in the row
                row = matching.iloc[0]
                email = None
                for val in row.values:
                    try:
                        s = str(val)
                        if "@" in s and "." in s:
                            email = s.strip()
                            break
                    except Exception:
                        continue
                if not email:
                    st.warning(f"No email for {name}")
                    continue

                pdf_bytes = _call_pdf_compat(generate_duty_pdf, name, schedule_df, staff_df, start_date, end_date, exam_type, college_logo_bytes, uni_logo_bytes, None)
                sent = send_email_with_attachment(email, f"Duty Allotment - {name}", "Please find attached your duty allotment.", pdf_bytes, f"Duty_{name}.pdf")
                if sent:
                    st.success(f"Email sent to {email}")
                else:
                    st.error(f"Failed to send to {email}")

st.header("Attendance Marking")
# Small status badge showing whether schedule was loaded or freshly generated
badge_text = None
badge_state = st.session_state.get('schedule_loaded_from')
badge_ts = st.session_state.get('schedule_loaded_timestamp')
if badge_state == 'disk':
    short_ts = badge_ts[:19] if badge_ts else 'unknown'
    badge_text = f"✅ Loaded from disk ({short_ts})"
elif badge_state == 'generated':
    short_ts = badge_ts[:19] if badge_ts else 'now'
    badge_text = f"🟢 Generated ({short_ts})"
else:
    badge_text = "⚠️ No schedule loaded"

colx, coly = st.columns([8, 2])
with colx:
    pass
with coly:
    st.markdown(f"**Status:** {badge_text}")

if "schedule_df" in st.session_state:
    # Normalize and validate schedule before use
    schedule_df = ensure_schedule_schema(st.session_state.get("schedule_df"))
    if schedule_df is None:
        st.error("Current schedule is malformed or missing required columns (date/session/assigned). Please restore a valid schedule or regenerate it.")
        st.stop()
    dates = sorted(schedule_df["date"].unique())
    st.write("Mark attendance date-wise and session-wise. Selected = present; unselected = absent.")
    # Load persisted attendance state if present
    if "attendance" not in st.session_state:
        st.session_state["attendance"] = {}
        if os.path.exists("attendance_state.json"):
            try:
                with open("attendance_state.json", "r", encoding="utf-8") as f:
                    st.session_state["attendance"] = json.load(f)
            except Exception:
                st.warning("Unable to load persisted attendance state; starting fresh.")

    for d in dates:
        st.subheader(d.strftime("%Y-%m-%d"))
        morning = schedule_df[(schedule_df["date"] == d) & (schedule_df["session"] == "Morning")]
        evening = schedule_df[(schedule_df["date"] == d) & (schedule_df["session"] == "Evening")]
        morning_assigned = morning.iloc[0]["assigned"] if not morning.empty else []
        evening_assigned = evening.iloc[0]["assigned"] if not evening.empty else []

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Morning (10.00 a.m. to 01.00 p.m.)**")
            present_m = []
            # show checkbox per assigned supervisor
            for idx, name in enumerate(sorted(morning_assigned)):
                safe_key = f"att_{d.strftime('%Y%m%d')}_m_{idx}_{name.replace(' ', '_')}"
                prev = False
                prev_info = st.session_state.get("attendance", {}).get(d.strftime("%Y-%m-%d"), {})
                if prev_info:
                    prev = name in prev_info.get("Morning_present", [])
                checked = st.checkbox(name, value=prev, key=safe_key)
                if checked:
                    present_m.append(name)
        with col2:
            st.markdown("**Evening (02.00 p.m. to 05.00 p.m.)**")
            present_e = []
            for idx, name in enumerate(sorted(evening_assigned)):
                safe_key = f"att_{d.strftime('%Y%m%d')}_e_{idx}_{name.replace(' ', '_')}"
                prev = False
                prev_info = st.session_state.get("attendance", {}).get(d.strftime("%Y-%m-%d"), {})
                if prev_info:
                    prev = name in prev_info.get("Evening_present", [])
                checked = st.checkbox(name, value=prev, key=safe_key)
                if checked:
                    present_e.append(name)

        # Persist attendance
        st.session_state["attendance"][d.strftime("%Y-%m-%d")] = {
            "Morning_present": present_m,
            "Morning_assigned": morning_assigned,
            "Evening_present": present_e,
            "Evening_assigned": evening_assigned,
        }
        # Auto-save attendance state to disk so refresh won't lose marks
        save_attendance_state(st.session_state["attendance"])

        # Per-date save + memo generation button
        save_key = f"save_{d.strftime('%Y%m%d')}"
        if st.button(f"Save & generate memos for {d.strftime('%Y-%m-%d')}", key=save_key):
            date_str = d.strftime("%Y-%m-%d")
            info = st.session_state["attendance"][date_str]
            # write per-date CSV
            per_rows = []
            for name in info["Morning_assigned"]:
                per_rows.append({"Date": date_str, "Session": "Morning", "Name": name, "Present": (name in info["Morning_present"])})
            for name in info["Evening_assigned"]:
                per_rows.append({"Date": date_str, "Session": "Evening", "Name": name, "Present": (name in info["Evening_present"])})
            per_df = pd.DataFrame(per_rows)
            try:
                per_df.to_csv(f"attendance_{date_str}.csv", index=False)
            except Exception:
                st.warning("Unable to write per-date CSV to disk; proceeding to generate memos in memory.")

            # Build absentees for this date only
            abs_map_date = {}
            for name in info["Morning_assigned"]:
                if name not in info["Morning_present"]:
                    abs_map_date.setdefault(name, []).append((d, "Morning"))
            for name in info["Evening_assigned"]:
                if name not in info["Evening_present"]:
                    abs_map_date.setdefault(name, []).append((d, "Evening"))

            if not abs_map_date:
                st.success(f"Attendance saved for {date_str}. No absentees found.")
            else:
                # Ensure global absentee map exists and merge
                st.session_state.setdefault("absentee_map", {})

                # Determine signature bytes (try uploaded memo signature first, then sign.jpg file)
                sign_bytes = None
                try:
                    uploaded = st.session_state.get("sign_upload")
                    if uploaded:
                        sign_bytes = uploaded.read()
                except Exception:
                    sign_bytes = None
                if not sign_bytes:
                    default_sign_path = os.path.join(os.getcwd(), "sign.jpg")
                    if os.path.exists(default_sign_path):
                        try:
                            with open(default_sign_path, "rb") as f:
                                sign_bytes = f.read()
                        except Exception:
                            sign_bytes = None

                # Generate memo PDFs for absent supervisors for this date and add to absentee_map
                generated = 0
                for name, absences in abs_map_date.items():
                    st.session_state["absentee_map"].setdefault(name, []).extend(absences)
                    try:
                        from pdf_utils import generate_absence_memo
                        memo_pdf = _call_memo_compat(generate_absence_memo, name, absences, staff_df, None, None, sign_bytes)
                    except Exception:
                        pdf_mod = importlib.import_module("pdf_utils")
                        memo_pdf = _call_memo_compat(pdf_mod.generate_absence_memo, name, absences, staff_df, None, None, sign_bytes)

                    # Save memo to file if possible and provide a download
                    fname = f"Memo_{name.replace(' ', '_')}_{date_str}.pdf"
                    try:
                        with open(fname, "wb") as f:
                            f.write(memo_pdf)
                    except Exception:
                        pass
                    st.download_button(f"Download memo for {name} ({date_str})", data=memo_pdf, file_name=fname, mime="application/pdf", key=f"download_memo_{date_str}_{name.replace(' ', '_')}")
                    generated += 1

                # Persist attendance state and notify
                save_attendance_state(st.session_state["attendance"])
                st.success(f"Attendance saved for {date_str} and memos generated for {generated} supervisor(s). They are also available under 'Absence Memos'.")

    if st.button("Save attendance"):
        # Persist to CSV: one row per date-session-supervisor with status
        rows = []
        for date_str, info in st.session_state["attendance"].items():
            for name in info["Morning_assigned"]:
                rows.append({"Date": date_str, "Session": "Morning", "Name": name, "Present": (name in info["Morning_present"])})
            for name in info["Evening_assigned"]:
                rows.append({"Date": date_str, "Session": "Evening", "Name": name, "Present": (name in info["Evening_present"])})
        df_att = pd.DataFrame(rows)
        df_att.to_csv("attendance_detailed.csv", index=False)
        # Also write per-date CSVs
        for date_str, info in st.session_state["attendance"].items():
            per_rows = []
            for name in info["Morning_assigned"]:
                per_rows.append({"Date": date_str, "Session": "Morning", "Name": name, "Present": (name in info["Morning_present"])})
            for name in info["Evening_assigned"]:
                per_rows.append({"Date": date_str, "Session": "Evening", "Name": name, "Present": (name in info["Evening_present"])})
            per_df = pd.DataFrame(per_rows)
            per_df.to_csv(f"attendance_{date_str}.csv", index=False)
        # Persist attendance state JSON as well
        save_attendance_state(st.session_state["attendance"])
        st.success("Attendance saved to attendance_detailed.csv and per-date files (attendance_YYYY-MM-DD.csv)")
        # Also provide an explicit download backup of current attendance state
        try:
            with open('attendance_state.json', 'rb') as f:
                b = f.read()
            st.sidebar.download_button('Download attendance backup (JSON)', data=b, file_name='attendance_state.json', mime='application/json')
        except Exception:
            pass

    # Provide consolidated download (horizontal) with date-session columns
    if 'attendance' in st.session_state and st.session_state['attendance']:
        def consolidated_attendance_excel_bytes(att_map):
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            wb = Workbook()
            ws = wb.active
            ws.title = 'Consolidated Attendance'

            # Collect all dates and names
            dates = sorted(att_map.keys())
            names = sorted({n for info in att_map.values() for n in info['Morning_assigned'] + info['Evening_assigned']})

            # Header
            headers = ['Name']
            for d in dates:
                headers.append(f"{d} Morning")
                headers.append(f"{d} Evening")
            headers.extend(['Total Assigned', 'Total Present', 'Total Absent'])
            for ci, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Rows per name
            for ri, name in enumerate(names, start=2):
                ws.cell(row=ri, column=1, value=name)
                total_assigned = 0
                total_present = 0
                col = 2
                for d in dates:
                    info = att_map.get(d, {})
                    m_assigned = name in info.get('Morning_assigned', [])
                    e_assigned = name in info.get('Evening_assigned', [])
                    if m_assigned:
                        total_assigned += 1
                        present = name in info.get('Morning_present', [])
                        if present:
                            total_present += 1
                        ws.cell(row=ri, column=col, value='P' if present else 'A')
                    else:
                        ws.cell(row=ri, column=col, value='')
                    col += 1
                    if e_assigned:
                        total_assigned += 1
                        present = name in info.get('Evening_present', [])
                        if present:
                            total_present += 1
                        ws.cell(row=ri, column=col, value='P' if present else 'A')
                    else:
                        ws.cell(row=ri, column=col, value='')
                    col += 1

                ws.cell(row=ri, column=col, value=total_assigned)
                ws.cell(row=ri, column=col+1, value=total_present)
                ws.cell(row=ri, column=col+2, value=(total_assigned - total_present))

            # Auto width
            for i in range(1, ws.max_column+1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio.read()

        consolidated_bytes = consolidated_attendance_excel_bytes(st.session_state['attendance'])
        st.download_button("Download consolidated attendance (Excel)", data=consolidated_bytes, file_name="Consolidated_Attendance.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_consolidated_attendance")
    st.markdown("---")
    st.subheader("Absence Memos")
    memo_subject = st.text_input("Memo email subject", value=st.session_state.get("memo_subject", "Absence from invigilation duty"))
    sign_file = st.file_uploader("Signature image (optional, used in memo and duty PDF)", type=["png","jpg","jpeg"], key="sign_upload")
    sign_bytes = sign_file.read() if sign_file else None

    if st.button("Generate memos for absentees"):
        # Build list of absentees per supervisor
        absentee_map = {}
        for date_str, info in st.session_state["attendance"].items():
            d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            for name in info["Morning_assigned"]:
                if name not in info["Morning_present"]:
                    absentee_map.setdefault(name, []).append((d, "Morning"))
            for name in info["Evening_assigned"]:
                if name not in info["Evening_present"]:
                    absentee_map.setdefault(name, []).append((d, "Evening"))

        if not absentee_map:
            st.success("No absentees found.")
        else:
            st.session_state["absentee_map"] = absentee_map
            st.success(f"Generated memos for {len(absentee_map)} absent supervisor(s). You can download or email them below.")

    # If memos exist, show downloads and email option
    if "absentee_map" in st.session_state:
        for name, absences in st.session_state["absentee_map"].items():
            memo_pdf = None
            try:
                pdf_mod = importlib.import_module("pdf_utils")
                memo_pdf = _call_memo_compat(pdf_mod.generate_absence_memo, name, absences, staff_df, None, None, sign_bytes)
            except Exception:
                try:
                    from pdf_utils import generate_absence_memo
                    memo_pdf = _call_memo_compat(generate_absence_memo, name, absences, staff_df, None, None, sign_bytes)
                except Exception:
                    memo_pdf = None
            st.download_button(f"Download memo for {name}", data=memo_pdf, file_name=f"Memo_{name}.pdf", mime="application/pdf", key=f"download_bulk_memo_{name.replace(' ', '_')}")

        memo_send_emails = st.multiselect("Select absentees to email memos", options=list(st.session_state["absentee_map"].keys()))
        memo_subject_input = st.text_input("Memo email subject (for sending)", value=memo_subject)
        if st.button("Send memo emails to selected"):
            for name in memo_send_emails:
                # find email
                matching = staff_df[staff_df.iloc[:,1].str.strip()==name]
                if matching.empty:
                    st.warning(f"No email for {name}")
                    continue
                # find email in row
                row = matching.iloc[0]
                email = None
                for val in row.values:
                    try:
                        s = str(val)
                        if "@" in s and "." in s:
                            email = s.strip()
                            break
                    except Exception:
                        continue
                if not email:
                    st.warning(f"No email for {name}")
                    continue
                # generate memo pdf bytes
                pdf_mod = importlib.import_module("pdf_utils")
                memo_pdf = _call_memo_compat(pdf_mod.generate_absence_memo, name, st.session_state["absentee_map"][name], staff_df, None, None, sign_bytes)
                sent = send_email_with_attachment(email, memo_subject_input, "Please find attached your absence memo.", memo_pdf, f"Memo_{name}.pdf")
                if sent:
                    st.success(f"Memo sent to {email}")
                else:
                    st.error(f"Failed to send memo to {email}")
else:
    st.info("Generate schedule to mark attendance")
